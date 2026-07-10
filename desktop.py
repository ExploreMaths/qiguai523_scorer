# -*- coding: utf-8 -*-
"""七怪五二三 计分器（两副牌）—— PyQt6 桌面版，复刻网页版。"""
import json
import os
import sys

from PyQt6.QtCore import Qt
from PyQt6.QtGui import QColor, QFont, QFontDatabase
from PyQt6.QtWidgets import (
    QApplication,
    QDialog,
    QFileDialog,
    QFrame,
    QGridLayout,
    QHBoxLayout,
    QHeaderView,
    QLabel,
    QLineEdit,
    QMainWindow,
    QPushButton,
    QScrollArea,
    QSizePolicy,
    QStackedWidget,
    QTableWidget,
    QTableWidgetItem,
    QVBoxLayout,
    QWidget,
)

BASE_DIR = os.path.dirname(os.path.abspath(__file__))
DATA_FILE = os.path.join(BASE_DIR, "qiguai523_data.json")

BOMB_TYPES = [
    ("＋3张 ×2", 2),
    ("＋4张 ×4", 4),
    ("＋5张 ×8", 8),
    ("＋6张 ×16", 16),
    ("＋7张 ×32", 32),
]

COLORS = {
    "bg": "#f8fafc",
    "card": "#ffffff",
    "border": "#e2e8f0",
    "text": "#1e293b",
    "text_secondary": "#64748b",
    "primary": "#3b82f6",
    "primary_hover": "#2563eb",
    "primary_light": "#eff6ff",
    "success_bg": "#d1fae5",
    "success_text": "#065f46",
    "error": "#ef4444",
    "error_bg": "#fee2e2",
    "error_text": "#991b1b",
    "warning_bg": "#fffbeb",
    "warning_text": "#92400e",
}

BTN_STYLE = {
    "primary": (
        "QPushButton{background:%s;color:#fff;border:none;border-radius:8px;"
        "padding:12px 20px;font-size:22px;font-weight:700;}"
        "QPushButton:hover{background:%s;}" % (COLORS["primary"], COLORS["primary_hover"])
    ),
    "outline": (
        "QPushButton{background:%s;color:%s;border:1px solid %s;border-radius:8px;"
        "padding:12px 20px;font-size:22px;font-weight:700;}"
        "QPushButton:hover{background:%s;}"
        % (COLORS["card"], COLORS["text"], COLORS["border"], COLORS["bg"])
    ),
    "danger": (
        "QPushButton{background:%s;color:%s;border:1px solid %s;border-radius:8px;"
        "padding:12px 20px;font-size:22px;font-weight:700;}"
        "QPushButton:hover{background:%s;}"
        % (COLORS["card"], COLORS["error"], COLORS["border"], COLORS["error_bg"])
    ),
    "bomb": (
        "QPushButton{background:%s;color:%s;border:1px solid #fcd34d;border-radius:8px;"
        "padding:12px 16px;font-size:22px;font-weight:700;}"
        "QPushButton:hover{background:#fef3c7;}"
        % (COLORS["warning_bg"], COLORS["warning_text"])
    ),
    "sm": (
        "QPushButton{background:%s;color:%s;border:1px solid %s;border-radius:8px;"
        "padding:8px 14px;font-size:18px;font-weight:700;}"
        "QPushButton:hover{background:%s;}"
        % (COLORS["card"], COLORS["text"], COLORS["border"], COLORS["bg"])
    ),
    "sm_danger": (
        "QPushButton{background:%s;color:%s;border:1px solid %s;border-radius:8px;"
        "padding:8px 14px;font-size:18px;font-weight:700;}"
        "QPushButton:hover{background:%s;}"
        % (COLORS["card"], COLORS["error"], COLORS["border"], COLORS["error_bg"])
    ),
    "icon": (
        "QPushButton{background:%s;color:%s;border:1px solid %s;border-radius:8px;"
        "padding:10px 12px;font-size:20px;font-weight:700;}"
        "QPushButton:hover{background:%s;}"
        % (COLORS["card"], COLORS["error"], COLORS["border"], COLORS["error_bg"])
    ),
}


def make_button(text, kind="outline"):
    b = QPushButton(text)
    b.setStyleSheet(BTN_STYLE[kind])
    b.setCursor(Qt.CursorShape.PointingHandCursor)
    return b


def make_card():
    f = QFrame()
    f.setObjectName("card")
    f.setStyleSheet(
        "QFrame#card{background:%s;border:1px solid %s;border-radius:8px;}"
        % (COLORS["card"], COLORS["border"])
    )
    return f


# --------------------------------------------------------------------------- #
# 弹窗
# --------------------------------------------------------------------------- #
class Modal(QDialog):
    def __init__(self, parent, message, buttons):
        super().__init__(parent)
        self.setWindowTitle("提示")
        self.setModal(True)
        self.setMinimumWidth(420)
        self.setStyleSheet(
            "QDialog{background:#fff;border-radius:8px;}"
            "QLabel{color:%s;font-size:24px;}" % COLORS["text"]
        )
        lay = QVBoxLayout(self)
        lay.setContentsMargins(32, 32, 32, 32)
        lay.setSpacing(24)
        msg = QLabel(message)
        msg.setWordWrap(True)
        msg.setAlignment(Qt.AlignmentFlag.AlignCenter)
        lay.addWidget(msg)

        row = QHBoxLayout()
        row.setSpacing(12)
        for text, kind, cb in buttons:
            b = make_button(text, kind)
            b.setMinimumWidth(90)
            b.clicked.connect(lambda _=False, c=cb: self._on_click(c))
            row.addWidget(b)
        lay.addLayout(row)
        self._callback = None

    def _on_click(self, cb):
        self._callback = cb
        self.accept()

    @staticmethod
    def alert(parent, message, on_ok=None):
        dlg = Modal(parent, message, [("确定", "primary", on_ok)])
        dlg.exec()
        if on_ok:
            on_ok()

    @staticmethod
    def confirm(parent, message, on_confirm=None, on_cancel=None):
        dlg = Modal(
            parent,
            message,
            [
                ("取消", "outline", on_cancel),
                ("确定", "primary", on_confirm),
            ],
        )
        dlg.exec()
        if dlg._callback is on_confirm and on_confirm:
            on_confirm()
        elif dlg._callback is on_cancel and on_cancel:
            on_cancel()


# --------------------------------------------------------------------------- #
# 状态持久化
# --------------------------------------------------------------------------- #
def default_state():
    return {
        "players": [],
        "rounds": [],
        "currentRoundIndex": 0,
        "presets": [],
        "defaultPreset": None,
        "currentView": "setup",
    }


def load_state():
    if not os.path.exists(DATA_FILE):
        return default_state()
    try:
        with open(DATA_FILE, "r", encoding="utf-8") as f:
            data = json.load(f)
        st = default_state()
        st.update({k: data.get(k, st[k]) for k in st})
        return st
    except Exception:
        return default_state()


def save_state(st):
    try:
        with open(DATA_FILE, "w", encoding="utf-8") as f:
            json.dump(st, f, ensure_ascii=False, indent=2)
    except Exception:
        pass


# --------------------------------------------------------------------------- #
# 设置玩家页
# --------------------------------------------------------------------------- #
class SetupPage(QWidget):
    def __init__(self, app):
        super().__init__()
        self.app = app
        self.player_inputs = []
        root = QVBoxLayout(self)
        root.setContentsMargins(0, 0, 0, 0)

        card = make_card()
        lay = QVBoxLayout(card)
        lay.setContentsMargins(24, 24, 24, 24)
        lay.setSpacing(14)

        self.title = QLabel("设置玩家")
        self.title.setAlignment(Qt.AlignmentFlag.AlignCenter)
        self.title.setStyleSheet(
            "font-size:39px;font-weight:700;color:%s;" % COLORS["text"]
        )
        lay.addWidget(self.title)

        hint = QLabel("每名玩家一行，可点击右侧 🗑 删除")
        hint.setAlignment(Qt.AlignmentFlag.AlignCenter)
        hint.setStyleSheet(
            "font-size:21px;color:%s;" % COLORS["text_secondary"]
        )
        lay.addWidget(hint)

        # 固定组合
        preset_box = QFrame()
        preset_box.setStyleSheet(
            "QFrame{background:%s;border:1px solid %s;border-radius:8px;}"
            % (COLORS["bg"], COLORS["border"])
        )
        pb = QVBoxLayout(preset_box)
        pb.setContentsMargins(16, 16, 16, 16)
        pb.setSpacing(10)
        pt = QLabel("固定组合")
        pt.setStyleSheet(
            "font-size:21px;font-weight:700;color:%s;" % COLORS["text_secondary"]
        )
        pb.addWidget(pt)
        self.preset_row = QHBoxLayout()
        self.preset_row.setSpacing(8)
        pb.addLayout(self.preset_row)
        add_row = QHBoxLayout()
        add_row.setSpacing(8)
        self.preset_name = QLineEdit()
        self.preset_name.setPlaceholderText("组合名称")
        self.preset_name.setStyleSheet(self._input_style(21))
        add_row.addWidget(self.preset_name, 1)
        save_btn = make_button("保存当前为组合", "outline")
        save_btn.clicked.connect(self.save_current_as_preset)
        add_row.addWidget(save_btn)
        pb.addLayout(add_row)
        lay.addWidget(preset_box)

        # 玩家列表
        self.players_box = QVBoxLayout()
        self.players_box.setSpacing(10)
        lay.addLayout(self.players_box)

        add_player = make_button("＋ 添加玩家", "outline")
        add_player.clicked.connect(lambda: self.add_player_row(""))
        lay.addWidget(add_player)

        start_btn = make_button("开始计分", "primary")
        start_btn.clicked.connect(self.app.start_game)
        lay.addWidget(start_btn)

        root.addWidget(card)

    @staticmethod
    def _input_style(size):
        return (
            "QLineEdit{padding:12px 14px;font-size:%dpx;border:1px solid %s;"
            "border-radius:8px;background:%s;color:%s;}"
            "QLineEdit:focus{border:2px solid %s;}"
            % (size, COLORS["border"], COLORS["bg"], COLORS["text"], COLORS["primary"])
        )

    def add_player_row(self, name=""):
        row = QHBoxLayout()
        row.setSpacing(10)
        inp = QLineEdit()
        inp.setPlaceholderText("玩家姓名")
        inp.setText(name)
        inp.setStyleSheet(self._input_style(22))
        row.addWidget(inp, 1)
        rm = make_button("🗑", "icon")
        rm.setFixedWidth(56)
        rm.clicked.connect(lambda _=False, i=inp: self.remove_player_row(i))
        row.addWidget(rm)
        container = QWidget()
        container.setLayout(row)
        self.players_box.addWidget(container)
        self.player_inputs.append((inp, container))

    def remove_player_row(self, inp):
        if len(self.player_inputs) <= 2:
            Modal.alert(self, "至少需要两名玩家")
            return
        for i, (ip, c) in enumerate(self.player_inputs):
            if ip is inp:
                c.setParent(None)
                c.deleteLater()
                self.player_inputs.pop(i)
                break

    def current_players(self):
        return [ip.text().strip() for ip, _ in self.player_inputs if ip.text().strip()]

    def rebuild(self, players):
        # 清空
        for _, c in self.player_inputs:
            c.setParent(None)
            c.deleteLater()
        self.player_inputs = []
        if not players:
            for i in range(1, 5):
                self.add_player_row("玩家%d" % i)
        else:
            for p in players:
                self.add_player_row(p)
        self.render_presets()

    def render_presets(self):
        # 清空 preset_row
        while self.preset_row.count():
            item = self.preset_row.takeAt(0)
            w = item.widget()
            if w:
                w.setParent(None)
                w.deleteLater()
        presets = self.app.state["presets"]
        if not presets:
            empty = QLabel("暂无固定组合")
            empty.setStyleSheet(
                "font-size:19px;color:%s;" % COLORS["text_secondary"]
            )
            self.preset_row.addWidget(empty)
            self.preset_row.addStretch(1)
            return
        for preset in presets:
            chip = self._make_chip(preset)
            self.preset_row.addWidget(chip)
        self.preset_row.addStretch(1)

    def _make_chip(self, preset):
        f = QFrame()
        is_default = self.app.state["defaultPreset"] == preset["name"]
        border = COLORS["primary"] if is_default else COLORS["border"]
        bg = COLORS["primary_light"] if is_default else COLORS["card"]
        f.setStyleSheet(
            "QFrame{background:%s;border:1px solid %s;border-radius:8px;}" % (bg, border)
        )
        h = QHBoxLayout(f)
        h.setContentsMargins(10, 6, 10, 6)
        h.setSpacing(6)
        name_btn = QPushButton(preset["name"])
        name_btn.setCursor(Qt.CursorShape.PointingHandCursor)
        name_btn.setStyleSheet(
            "QPushButton{background:transparent;border:none;font-size:19px;"
            "color:%s;font-weight:%d;}"
            % (COLORS["primary"] if is_default else COLORS["text"], 700 if is_default else 400)
        )
        name_btn.clicked.connect(lambda _=False, n=preset["name"]: self.app.apply_preset(n))
        h.addWidget(name_btn)
        if is_default:
            star = QLabel("★")
            star.setStyleSheet("color:%s;font-size:18px;" % "#f59e0b")
            h.addWidget(star)
        set_def = QPushButton("★")
        set_def.setToolTip("设为默认")
        set_def.setCursor(Qt.CursorShape.PointingHandCursor)
        set_def.setStyleSheet(
            "QPushButton{background:transparent;border:none;color:%s;font-size:18px;}"
            % COLORS["text_secondary"]
        )
        set_def.clicked.connect(lambda _=False, n=preset["name"]: self.app.set_default_preset(n))
        h.addWidget(set_def)
        dele = QPushButton("🗑")
        dele.setToolTip("删除")
        dele.setCursor(Qt.CursorShape.PointingHandCursor)
        dele.setStyleSheet(
            "QPushButton{background:transparent;border:none;color:%s;font-size:18px;}"
            % COLORS["text_secondary"]
        )
        dele.clicked.connect(lambda _=False, n=preset["name"]: self.app.delete_preset(n))
        h.addWidget(dele)
        return f

    def save_current_as_preset(self):
        name = self.preset_name.text().strip()
        if not name:
            Modal.alert(self, "请输入组合名称")
            return
        players = self.current_players()
        if len(players) < 2:
            Modal.alert(self, "至少需要两名玩家才能保存组合")
            return
        presets = self.app.state["presets"]
        exists = any(p["name"] == name for p in presets)

        def do_save():
            for p in presets:
                if p["name"] == name:
                    p["players"] = players
                    break
            else:
                presets.append({"name": name, "players": players})
            self.preset_name.clear()
            self.app.save()
            self.render_presets()

        if exists:
            Modal.confirm(self, '组合 "%s" 已存在，是否覆盖？' % name, do_save)
        else:
            do_save()


# --------------------------------------------------------------------------- #
# 计分页
# --------------------------------------------------------------------------- #
class GamePage(QWidget):
    def __init__(self, app):
        super().__init__()
        self.app = app
        self.inputs = []
        self.final_labels = []
        self._built = False
        root = QVBoxLayout(self)
        root.setContentsMargins(0, 0, 0, 0)
        root.setSpacing(20)

        # 头部
        header = QHBoxLayout()
        self.h1 = QLabel("七怪五二三 计分器")
        self.h1.setStyleSheet(
            "font-size:42px;font-weight:700;color:%s;" % COLORS["text"]
        )
        header.addWidget(self.h1)
        header.addStretch(1)
        self.setup_btn = make_button("设置玩家", "sm")
        self.setup_btn.clicked.connect(self.app.open_setup)
        header.addWidget(self.setup_btn)
        self.del_btn = make_button("删除该局", "sm_danger")
        self.del_btn.clicked.connect(self.app.delete_current_round)
        header.addWidget(self.del_btn)
        self.round_badge = QLabel("第 1 局 / 共 1 局")
        self.round_badge.setStyleSheet(
            "QLabel{background:%s;border:1px solid %s;color:%s;border-radius:8px;"
            "padding:8px 18px;font-size:21px;font-weight:700;}"
            % (COLORS["card"], COLORS["border"], COLORS["primary"])
        )
        header.addWidget(self.round_badge)
        root.addLayout(header)

        # 计分卡
        self.score_card = make_card()
        self.score_layout = QVBoxLayout(self.score_card)
        self.score_layout.setContentsMargins(24, 24, 24, 24)
        self.score_layout.setSpacing(24)
        root.addWidget(self.score_card)

        # 本局合计
        total_card = make_card()
        tc = QHBoxLayout(total_card)
        tc.setContentsMargins(24, 18, 24, 18)
        tc.addStretch(1)
        lbl = QLabel("本局合计：")
        lbl.setStyleSheet("font-size:27px;color:%s;" % COLORS["text"])
        tc.addWidget(lbl)
        self.round_total = QLabel("0")
        self.round_total.setStyleSheet(
            "font-size:33px;font-weight:700;color:%s;" % COLORS["text"]
        )
        tc.addWidget(self.round_total)
        tc.addStretch(1)
        root.addWidget(total_card)

        # 操作
        act = make_card()
        al = QVBoxLayout(act)
        al.setContentsMargins(24, 24, 24, 24)
        al.setSpacing(14)
        nav = QHBoxLayout()
        nav.setSpacing(12)
        self.prev_btn = make_button("← 上一局", "outline")
        self.prev_btn.clicked.connect(self.app.prev_round)
        nav.addWidget(self.prev_btn)
        self.next_btn = make_button("下一局 →", "outline")
        self.next_btn.clicked.connect(self.app.next_round)
        nav.addWidget(self.next_btn)
        al.addLayout(nav)
        finish = make_button("完成并查看完整表格", "primary")
        finish.clicked.connect(self.app.show_summary)
        al.addWidget(finish)
        root.addWidget(act)

    def build(self):
        # 清空 score_layout
        while self.score_layout.count():
            item = self.score_layout.takeAt(0)
            w = item.widget()
            if w:
                w.setParent(None)
                w.deleteLater()
            elif item.layout():
                self._clear_layout(item.layout())
        self.inputs = []
        self.final_labels = []
        players = self.app.state["players"]
        n = len(players)

        grid_host = QWidget()
        grid = QGridLayout(grid_host)
        grid.setHorizontalSpacing(12)
        grid.setVerticalSpacing(12)
        grid.setContentsMargins(0, 0, 0, 0)
        grid.setColumnMinimumWidth(0, 150)
        grid.setColumnStretch(0, 0)
        for c in range(1, n + 2):
            grid.setColumnStretch(c, 1)

        # 表头
        corner = QLabel("")
        grid.addWidget(corner, 0, 0)
        for i, p in enumerate(players):
            h = QLabel(p)
            h.setAlignment(Qt.AlignmentFlag.AlignCenter)
            h.setStyleSheet(
                "font-size:19px;font-weight:700;color:%s;" % COLORS["text_secondary"]
            )
            grid.addWidget(h, 0, i + 1)
        grid.addWidget(QLabel(""), 0, n + 1)

        # 原分数行
        lab1 = QLabel("原分数")
        lab1.setStyleSheet(
            "font-size:21px;font-weight:500;color:%s;" % COLORS["text"]
        )
        grid.addWidget(lab1, 1, 0)
        for i in range(n):
            inp = QLineEdit("0")
            inp.setAlignment(Qt.AlignmentFlag.AlignCenter)
            inp.setStyleSheet(
                "QLineEdit{padding:14px 8px;font-size:27px;text-align:center;"
                "border:1px solid %s;border-radius:8px;background:%s;color:%s;}"
                "QLineEdit:focus{border:2px solid %s;background:%s;}"
                % (COLORS["border"], COLORS["bg"], COLORS["text"], COLORS["primary"], COLORS["card"])
            )
            inp.textChanged.connect(self.app.update_display)
            grid.addWidget(inp, 1, i + 1)
            self.inputs.append(inp)

        # 翻倍后分数行
        lab2 = QLabel("翻倍后分数")
        lab2.setStyleSheet(
            "font-size:21px;font-weight:500;color:%s;" % COLORS["text"]
        )
        grid.addWidget(lab2, 2, 0)
        for i in range(n):
            fl = QLabel("0")
            fl.setAlignment(Qt.AlignmentFlag.AlignCenter)
            fl.setMinimumHeight(52)
            self._set_final_style(fl, "zero")
            grid.addWidget(fl, 2, i + 1)
            self.final_labels.append(fl)

        # 倍数框（跨两行）
        mul = QFrame()
        mul.setStyleSheet(
            "QFrame{border-left:1px solid %s;}" % COLORS["border"]
        )
        ml = QVBoxLayout(mul)
        ml.setContentsMargins(8, 8, 8, 8)
        ml.setSpacing(10)
        ml.addStretch(1)
        self.mul_value = QLabel("×1")
        self.mul_value.setAlignment(Qt.AlignmentFlag.AlignCenter)
        self.mul_value.setStyleSheet(
            "font-size:33px;font-weight:800;color:%s;" % COLORS["primary"]
        )
        ml.addWidget(self.mul_value)
        mr = QHBoxLayout()
        mr.setSpacing(8)
        undo = make_button("撤销", "sm")
        undo.clicked.connect(self.app.undo_multiplier)
        mr.addWidget(undo)
        reset = make_button("重置", "sm_danger")
        reset.clicked.connect(self.app.reset_multiplier)
        mr.addWidget(reset)
        ml.addLayout(mr)
        ml.addStretch(1)
        grid.addWidget(mul, 1, n + 1, 2, 1)

        self.score_layout.addWidget(grid_host)

        # 炸弹区
        sep = QFrame()
        sep.setFrameShape(QFrame.Shape.HLine)
        sep.setStyleSheet("color:%s;" % COLORS["border"])
        self.score_layout.addWidget(sep)
        bomb_row = QHBoxLayout()
        bomb_row.setSpacing(10)
        for label, factor in BOMB_TYPES:
            b = make_button(label, "bomb")
            b.clicked.connect(lambda _=False, f=factor: self.app.add_bomb(f))
            bomb_row.addWidget(b)
        self.score_layout.addLayout(bomb_row)
        self._built = True

    @staticmethod
    def _clear_layout(layout):
        while layout.count():
            item = layout.takeAt(0)
            w = item.widget()
            if w:
                w.setParent(None)
                w.deleteLater()
            elif item.layout():
                GamePage._clear_layout(item.layout())

    def _set_final_style(self, label, kind):
        if kind == "positive":
            bg, fg = COLORS["success_bg"], COLORS["success_text"]
        elif kind == "negative":
            bg, fg = COLORS["error_bg"], COLORS["error_text"]
        else:
            bg, fg = COLORS["bg"], COLORS["text_secondary"]
        label.setStyleSheet(
            "QLabel{background:%s;color:%s;border-radius:8px;font-size:27px;"
            "font-weight:700;padding:14px 8px;}" % (bg, fg)
        )

    def load_round(self, index):
        st = self.app.state
        st["currentRoundIndex"] = index
        rd = st["rounds"][index]
        self.round_badge.setText(
            "第 %d 局 / 共 %d 局" % (index + 1, len(st["rounds"]))
        )
        for i, inp in enumerate(self.inputs):
            inp.blockSignals(True)
            inp.setText(str(rd["scores"][i] if i < len(rd["scores"]) else 0))
            inp.blockSignals(False)
        self.app.update_display()
        self.prev_btn.setDisabled(index == 0)
        self.next_btn.setText(
            "＋ 新一局" if index == len(st["rounds"]) - 1 else "下一局 →"
        )


# --------------------------------------------------------------------------- #
# 汇总页
# --------------------------------------------------------------------------- #
class SummaryPage(QWidget):
    def __init__(self, app):
        super().__init__()
        self.app = app
        root = QVBoxLayout(self)
        root.setContentsMargins(0, 0, 0, 0)
        root.setSpacing(20)

        header = QHBoxLayout()
        h1 = QLabel("完整计分表")
        h1.setStyleSheet(
            "font-size:42px;font-weight:700;color:%s;" % COLORS["text"]
        )
        header.addWidget(h1)
        header.addStretch(1)
        root.addLayout(header)

        card = make_card()
        cl = QVBoxLayout(card)
        cl.setContentsMargins(0, 0, 0, 0)
        self.table = QTableWidget()
        self.table.setStyleSheet(
            "QTableWidget{background:#fff;border:none;font-size:22px;}"
            "QTableWidget::item{padding:12px 6px;}"
        )
        self.table.horizontalHeader().hide()
        self.table.verticalHeader().hide()
        self.table.setShowGrid(True)
        self.table.setEditTriggers(QTableWidget.EditTrigger.NoEditTriggers)
        self.table.setSelectionMode(QTableWidget.SelectionMode.NoSelection)
        self.table.setSizePolicy(QSizePolicy.Policy.Expanding, QSizePolicy.Policy.Expanding)
        cl.addWidget(self.table)
        root.addWidget(card, 1)

        act = make_card()
        al = QHBoxLayout(act)
        al.setContentsMargins(24, 24, 24, 24)
        al.setSpacing(12)
        back = make_button("返回计分", "outline")
        back.clicked.connect(self.app.back_to_game)
        al.addWidget(back)
        ex = make_button("导出 Excel", "primary")
        ex.clicked.connect(self.app.export_excel)
        al.addWidget(ex)
        im = make_button("导出图片", "primary")
        im.clicked.connect(self.app.export_image)
        al.addWidget(im)
        rs = make_button("重新开始", "danger")
        rs.clicked.connect(self.app.restart_game)
        al.addWidget(rs)
        root.addWidget(act)

    def _item(self, text, bold=False, size=22, color=None, bg=None):
        it = QTableWidgetItem(str(text))
        it.setTextAlignment(Qt.AlignmentFlag.AlignCenter)
        f = QFont(self.app.app_font_family)
        f.setPointSize(size)
        f.setBold(bold)
        it.setFont(f)
        if color:
            it.setForeground(QColor(color))
        if bg:
            it.setBackground(QColor(bg))
        else:
            it.setBackground(QColor("#ffffff"))
        return it

    def build(self):
        st = self.app.state
        players = st["players"]
        rounds = st["rounds"]
        n = len(players)
        col_count = 3 + n
        row_count = 1 + len(rounds) * 2 + 1
        t = self.table
        t.clear()
        t.setColumnCount(col_count)
        t.setRowCount(row_count)

        header = t.horizontalHeader()
        header.setSectionResizeMode(0, QHeaderView.ResizeMode.Fixed)
        header.setSectionResizeMode(1, QHeaderView.ResizeMode.Fixed)
        t.setColumnWidth(0, 48)
        t.setColumnWidth(1, 150)
        for c in range(2, col_count):
            header.setSectionResizeMode(c, QHeaderView.ResizeMode.Stretch)

        # 表头
        t.setItem(0, 0, self._item("", bold=True, size=22))
        t.setSpan(0, 0, 1, 2)
        for i, p in enumerate(players):
            t.setItem(0, i + 2, self._item(p, bold=True, size=22))
        t.setItem(0, col_count - 1, self._item("倍数", bold=True, size=22))

        totals = [0] * n
        r = 1
        for ri, rd in enumerate(rounds):
            # 原分数行
            t.setItem(r, 0, self._item(ri + 1, bold=True, size=27, color="#b45309"))
            t.setSpan(r, 0, 2, 1)
            t.setItem(r, 1, self._item("原分数", size=22, color=COLORS["text_secondary"]))
            for j, score in enumerate(rd["scores"]):
                t.setItem(r, j + 2, self._item(score, size=22))
            t.setItem(
                r, col_count - 1,
                self._item("×%s" % rd["multiplier"], bold=True, size=27, color="#1d4ed8"),
            )
            t.setSpan(r, col_count - 1, 2, 1)
            # 翻倍后分数行
            t.setItem(r + 1, 1, self._item("翻倍后分数", size=22, color=COLORS["text_secondary"]))
            for j, score in enumerate(rd["scores"]):
                final = score * rd["multiplier"]
                totals[j] += final
                color = "#047857" if final > 0 else "#b91c1c" if final < 0 else COLORS["text_secondary"]
                t.setItem(r + 1, j + 2, self._item(final, bold=True, size=24, color=color))
            r += 2

        # 总计行
        t.setItem(r, 0, self._item("总计", bold=True, size=27))
        t.setSpan(r, 0, 1, 2)
        for j, total in enumerate(totals):
            if total > 0:
                bg, fg = COLORS["success_bg"], COLORS["success_text"]
            elif total < 0:
                bg, fg = COLORS["error_bg"], COLORS["error_text"]
            else:
                bg, fg = COLORS["bg"], COLORS["text_secondary"]
            t.setItem(r, j + 2, self._item(total, bold=True, size=27, color=fg, bg=bg))
        t.setItem(r, col_count - 1, self._item("", bold=True, size=27))
        t.resizeRowsToContents()


# --------------------------------------------------------------------------- #
# 主窗口
# --------------------------------------------------------------------------- #
class MainWindow(QMainWindow):
    def __init__(self):
        super().__init__()
        self.setWindowTitle("七怪五二三 计分器")
        self.resize(1100, 860)
        self.state = load_state()

        self.app_font_family = self._load_fonts()
        self.setStyleSheet(
            "QMainWindow,QWidget{background:%s;color:%s;}"
            % (COLORS["bg"], COLORS["text"])
        )

        outer = QWidget()
        outer_lay = QVBoxLayout(outer)
        outer_lay.setContentsMargins(32, 32, 32, 32)
        outer_lay.setSpacing(0)

        self.stack = QStackedWidget()
        self.setup_page = SetupPage(self)
        self.game_page = GamePage(self)
        self.summary_page = SummaryPage(self)
        self.stack.addWidget(self.setup_page)
        self.stack.addWidget(self.game_page)
        self.stack.addWidget(self.summary_page)

        scroll = QScrollArea()
        scroll.setWidgetResizable(True)
        scroll.setFrameShape(QFrame.Shape.NoFrame)
        scroll.setWidget(self.stack)
        outer_lay.addWidget(scroll)
        self.setCentralWidget(outer)

        self._init_from_state()

    @staticmethod
    def _load_fonts():
        family = None
        for fn in ("LMROMAN12-REGULAR.OTF", "LMROMAN12-BOLD.OTF"):
            path = os.path.join(BASE_DIR, fn)
            if os.path.exists(path):
                fid = QFontDatabase.addApplicationFont(path)
                fams = QFontDatabase.applicationFontFamilies(fid)
                if fams and family is None:
                    family = fams[0]
        if family:
            app = QApplication.instance()
            f = QFont(family)
            f.setPointSize(15)
            app.setFont(f)
        return family or "Sans Serif"

    # ---- 视图 ----
    def switch_view(self, name):
        self.state["currentView"] = name
        if name == "setup":
            self.stack.setCurrentWidget(self.setup_page)
        elif name == "game":
            self.stack.setCurrentWidget(self.game_page)
        else:
            self.stack.setCurrentWidget(self.summary_page)
        self.save()

    def save(self):
        save_state(self.state)

    def _init_from_state(self):
        st = self.state
        if st["players"] and len(st["players"]) >= 2:
            if not st["rounds"]:
                st["rounds"] = [
                    {"scores": [0] * len(st["players"]), "multiplier": 1, "factors": []}
                ]
            self.game_page.build()
            if st["currentView"] == "setup":
                self.setup_page.rebuild(st["players"])
                self.switch_view("setup")
            elif st["currentView"] == "summary":
                self.summary_page.build()
                self.switch_view("summary")
            else:
                self.switch_view("game")
                self.game_page.load_round(st["currentRoundIndex"])
        else:
            if st["defaultPreset"]:
                preset = next(
                    (p for p in st["presets"] if p["name"] == st["defaultPreset"]), None
                )
                if preset and len(preset["players"]) >= 2:
                    st["players"] = list(preset["players"])
            st["currentView"] = "setup"
            self.setup_page.rebuild(st["players"])
            self.switch_view("setup")

    # ---- 组合 ----
    def apply_preset(self, name):
        preset = next((p for p in self.state["presets"] if p["name"] == name), None)
        if not preset or len(preset["players"]) < 2:
            return
        self.state["players"] = list(preset["players"])
        self.state["rounds"] = []
        self.state["currentRoundIndex"] = 0
        self.setup_page.rebuild(self.state["players"])
        self.save()

    def set_default_preset(self, name):
        self.state["defaultPreset"] = name
        self.save()
        self.setup_page.render_presets()

    def delete_preset(self, name):
        def do():
            self.state["presets"] = [
                p for p in self.state["presets"] if p["name"] != name
            ]
            if self.state["defaultPreset"] == name:
                self.state["defaultPreset"] = (
                    self.state["presets"][0]["name"] if self.state["presets"] else None
                )
            self.save()
            self.setup_page.render_presets()

        Modal.confirm(self, '确定删除组合 "%s" 吗？' % name, do)

    # ---- 设置玩家 / 开始 ----
    def open_setup(self):
        def do():
            self.state["players"] = []
            self.state["rounds"] = []
            self.state["currentRoundIndex"] = 0
            self.save()
            self.setup_page.rebuild([])
            self.switch_view("setup")

        Modal.confirm(self, "重新设置玩家将清除当前所有对局数据，是否继续？", do)

    def start_game(self):
        players = self.setup_page.current_players()
        if len(players) < 2:
            Modal.alert(self, "至少需要两名玩家")
            return
        if len(set(players)) != len(players):
            Modal.alert(self, "玩家姓名不能重复")
            return
        st = self.state
        old_len = len(st["players"])
        if old_len > 0 and len(players) != old_len:
            for rd in st["rounds"]:
                if len(players) > old_len:
                    rd["scores"] += [0] * (len(players) - old_len)
                else:
                    rd["scores"] = rd["scores"][: len(players)]
        st["players"] = players
        if not st["rounds"]:
            st["rounds"] = [
                {"scores": [0] * len(players), "multiplier": 1, "factors": []}
            ]
            st["currentRoundIndex"] = 0
        self.save()
        self.game_page.build()
        self.switch_view("game")
        self.game_page.load_round(st["currentRoundIndex"])

    def restart_game(self):
        def do():
            self.state["players"] = []
            self.state["rounds"] = []
            self.state["currentRoundIndex"] = 0
            self.save()
            self.setup_page.rebuild([])
            self.switch_view("setup")

        Modal.confirm(self, "确定要重新开始吗？当前所有数据将被清空。", do)

    # ---- 局 / 分数 ----
    def delete_current_round(self):
        st = self.state
        if len(st["rounds"]) <= 1:
            Modal.alert(self, "至少保留一局")
            return
        idx = st["currentRoundIndex"]

        def do():
            st["rounds"].pop(idx)
            if st["currentRoundIndex"] >= len(st["rounds"]):
                st["currentRoundIndex"] = len(st["rounds"]) - 1
            self.save()
            self.game_page.load_round(st["currentRoundIndex"])

        Modal.confirm(self, "确定删除第 %d 局吗？" % (idx + 1), do)

    def update_display(self):
        st = self.state
        if not st["rounds"]:
            return
        rd = st["rounds"][st["currentRoundIndex"]]
        total = 0
        for i, inp in enumerate(self.game_page.inputs):
            try:
                val = int(inp.text())
            except ValueError:
                val = 0
            if i >= len(rd["scores"]):
                rd["scores"].append(0)
            rd["scores"][i] = val
            final = val * rd["multiplier"]
            total += final
            fl = self.game_page.final_labels[i]
            fl.setText(str(final))
            kind = "positive" if final > 0 else "negative" if final < 0 else "zero"
            self.game_page._set_final_style(fl, kind)
        self.game_page.mul_value.setText("×%s" % rd["multiplier"])
        self.game_page.round_total.setText(str(total))
        self.save()

    def add_bomb(self, factor):
        rd = self.state["rounds"][self.state["currentRoundIndex"]]
        rd["multiplier"] *= factor
        rd["factors"].append(factor)
        self.update_display()

    def undo_multiplier(self):
        rd = self.state["rounds"][self.state["currentRoundIndex"]]
        if rd["factors"]:
            last = rd["factors"].pop()
            rd["multiplier"] = max(1, rd["multiplier"] // last)
        self.update_display()

    def reset_multiplier(self):
        rd = self.state["rounds"][self.state["currentRoundIndex"]]
        rd["multiplier"] = 1
        rd["factors"] = []
        self.update_display()

    def prev_round(self):
        if self.state["currentRoundIndex"] > 0:
            self.game_page.load_round(self.state["currentRoundIndex"] - 1)

    def next_round(self):
        st = self.state
        if not self.game_page._built:
            self.game_page.build()
        if st["currentRoundIndex"] >= len(st["rounds"]) - 1:
            st["rounds"].append(
                {"scores": [0] * len(st["players"]), "multiplier": 1, "factors": []}
            )
            self.game_page.load_round(len(st["rounds"]) - 1)
        else:
            self.game_page.load_round(st["currentRoundIndex"] + 1)

    # ---- 汇总 ----
    @staticmethod
    def _is_blank(rd):
        return all(s == 0 for s in rd["scores"]) and rd["multiplier"] == 1

    def _has_blank(self):
        return any(self._is_blank(rd) for rd in self.state["rounds"])

    def _remove_blank(self):
        st = self.state
        st["rounds"] = [rd for rd in st["rounds"] if not self._is_blank(rd)]
        if not st["rounds"]:
            st["rounds"] = [
                {"scores": [0] * len(st["players"]), "multiplier": 1, "factors": []}
            ]
        if st["currentRoundIndex"] >= len(st["rounds"]):
            st["currentRoundIndex"] = len(st["rounds"]) - 1
        self.save()

    def show_summary(self):
        self.update_display()

        def go():
            self.summary_page.build()
            self.switch_view("summary")

        if not self._has_blank():
            go()
            return
        Modal.confirm(
            self,
            "是否去除空白对局？",
            on_confirm=lambda: (self._remove_blank(), go()),
            on_cancel=go,
        )

    def back_to_game(self):
        if not self.game_page._built:
            self.game_page.build()
        self.switch_view("game")
        self.game_page.load_round(self.state["currentRoundIndex"])

    # ---- 导出 ----
    def export_excel(self):
        try:
            import openpyxl
            from openpyxl.styles import Alignment, Font, PatternFill
            from openpyxl.utils import get_column_letter
        except Exception:
            Modal.alert(self, "未安装 openpyxl，请先 pip install openpyxl")
            return

        st = self.state
        players = st["players"]
        rounds = st["rounds"]
        n = len(players)
        col_count = 3 + n

        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "计分表"
        align = Alignment(horizontal="center", vertical="center")
        FONT_NAME = "华文中宋"

        def font(sz, bold, color=None):
            return Font(name=FONT_NAME, size=sz, bold=bold, color=color)

        def setc(r, c, v, f, bg=None):
            cell = ws.cell(row=r, column=c, value=v)
            cell.font = f
            cell.alignment = align
            if bg:
                cell.fill = PatternFill("solid", fgColor=bg)
            return cell

        header_font = font(15, True, "1E293B")
        round_font = font(18, True, "B45309")
        label_font = font(15, False, "64748B")
        score_font = font(15, False, "1E293B")
        mul_font = font(18, True, "1D4ED8")
        total_label_font = font(18, True, "1E293B")
        total_value_font = font(16, True, "1E293B")

        # 表头
        setc(1, 1, "", header_font)
        ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=2)
        for i, p in enumerate(players):
            setc(1, i + 3, p, header_font)
        setc(1, col_count, "倍数", header_font)

        totals = [0] * n
        r = 2
        for ri, rd in enumerate(rounds):
            setc(r, 1, ri + 1, round_font)
            ws.merge_cells(start_row=r, start_column=1, end_row=r + 1, end_column=1)
            setc(r, 2, "原分数", label_font)
            for j, score in enumerate(rd["scores"]):
                setc(r, j + 3, score, score_font)
            setc(r, col_count, "×%s" % rd["multiplier"], mul_font)
            ws.merge_cells(
                start_row=r, start_column=col_count, end_row=r + 1, end_column=col_count
            )
            setc(r + 1, 2, "翻倍后分数", label_font)
            for j, score in enumerate(rd["scores"]):
                final = score * rd["multiplier"]
                totals[j] += final
                color = "047857" if final > 0 else "B91C1C" if final < 0 else "64748B"
                setc(r + 1, j + 3, final, font(16, True, color))
            r += 2

        # 总计
        setc(r, 1, "总计", total_label_font)
        ws.merge_cells(start_row=r, start_column=1, end_row=r, end_column=2)
        for j, total in enumerate(totals):
            if total > 0:
                bg, fg = "D1FAE5", "065F46"
            elif total < 0:
                bg, fg = "FEE2E2", "991B1B"
            else:
                bg, fg = "F1F5F9", "64748B"
            setc(r, j + 3, total, Font(name=FONT_NAME, size=16, bold=True, color=fg), bg)
        setc(r, col_count, "", total_value_font)

        # 列宽
        ws.column_dimensions[get_column_letter(1)].width = 7
        ws.column_dimensions[get_column_letter(2)].width = 20
        for c in range(3, col_count + 1):
            ws.column_dimensions[get_column_letter(c)].width = 16
        # 行高
        for rr in range(1, r + 1):
            ws.row_dimensions[rr].height = 33

        path, _ = QFileDialog.getSaveFileName(
            self, "导出 Excel", os.path.join(BASE_DIR, "七怪五二三计分表.xlsx"),
            "Excel 文件 (*.xlsx)",
        )
        if path:
            if not path.lower().endswith(".xlsx"):
                path += ".xlsx"
            wb.save(path)
            Modal.alert(self, "已导出：\n%s" % path)

    def export_image(self):
        path, _ = QFileDialog.getSaveFileName(
            self, "导出图片", os.path.join(BASE_DIR, "七怪五二三计分表.png"),
            "PNG 图片 (*.png)",
        )
        if not path:
            return
        if not path.lower().endswith(".png"):
            path += ".png"
        pix = self.summary_page.table.grab()
        if pix.save(path, "PNG"):
            Modal.alert(self, "已导出：\n%s" % path)
        else:
            Modal.alert(self, "图片导出失败，请重试。")


def main():
    app = QApplication(sys.argv)
    app.setApplicationName("七怪五二三计分器")
    w = MainWindow()
    w.show()
    sys.exit(app.exec())


if __name__ == "__main__":
    main()
